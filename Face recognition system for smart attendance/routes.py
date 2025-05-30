from app import app, db, Staff, Student, Attendance
from flask import render_template, request, redirect, url_for, flash, jsonify, send_file, session
from flask_login import login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from models.face_recognition_model import FaceRecognitionModel
import os
import cv2
import numpy as np
from datetime import datetime
import pandas as pd
import tempfile
import time
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from mtcnn import MTCNN
import tensorflow as tf

# Initialize MTCNN for faster face detection
detector = MTCNN()

# Enable GPU memory growth to prevent OOM errors
gpu_devices = tf.config.experimental.list_physical_devices('GPU')
for device in gpu_devices:
    tf.config.experimental.set_memory_growth(device, True)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        # Check hardcoded credentials
        if username == 'TEAM-5' and password == 'MLF-PROJECT':
            session['logged_in'] = True
            flash('Login successful!', 'success')
            return redirect(url_for('dashboard'))
        flash('Invalid username or password', 'danger')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.pop('logged_in', None)
    flash('Logged out successfully', 'success')
    return redirect(url_for('index'))

@app.route('/dashboard')
def dashboard():
    if not session.get('logged_in'):
        flash('Please login first', 'warning')
        return redirect(url_for('login'))
    
    # Get all students with their attendance records
    students = Student.query.all()
    
    # Convert students to JSON-serializable format with attendance data including timestamps
    students_data = []
    
    # Department-wise student data
    department_data = {}
    
    for student in students:
        attendance_records = Attendance.query.filter_by(student_id=student.id).all()
        attendance_data = [{
            'date': record.date.strftime('%Y-%m-%d'),
            'time': record.time.strftime('%H:%M:%S') if record.time else None,
            'status': record.status
        } for record in attendance_records]
        
        # Calculate attendance percentage for this student
        total_days = Attendance.query.with_entities(Attendance.date).distinct().count()
        student_present_days = Attendance.query.filter_by(
            student_id=student.id, 
            status='present'
        ).count()
        
        attendance_percentage = 0
        if total_days > 0:
            attendance_percentage = (student_present_days / total_days) * 100
        
        # Add attendance_percentage directly to the student object
        student.attendance_percentage = attendance_percentage
        student.attendance = attendance_data
        
        student_info = {
            'id': student.id,
            'register_no': student.register_no,
            'name': student.name,
            'batch': student.batch,
            'department': student.department,
            'attendance': attendance_data,
            'attendance_percentage': attendance_percentage
        }
        
        students_data.append(student_info)
        
        # Group by department for department-wise data
        if student.department not in department_data:
            department_data[student.department] = {
                'student_count': 0,
                'total_attendance_percentage': 0
            }
        
        department_data[student.department]['student_count'] += 1
        department_data[student.department]['total_attendance_percentage'] += attendance_percentage
    
    # Calculate average attendance percentage for each department
    for dept in department_data:
        if department_data[dept]['student_count'] > 0:
            department_data[dept]['avg_attendance'] = department_data[dept]['total_attendance_percentage'] / department_data[dept]['student_count']
        else:
            department_data[dept]['avg_attendance'] = 0
    
    # Get today's date for attendance status
    today = datetime.now().date().strftime('%Y-%m-%d')
    
    return render_template('dashboard.html', students=students, students_data=students_data, 
                           department_data=department_data, today=today)

@app.route('/register_student', methods=['GET', 'POST'])
def register_student():
    if not session.get('logged_in'):
        flash('Please login first', 'warning')
        return redirect(url_for('login'))
    if request.method == 'POST':
        register_no = request.form.get('register_no')
        name = request.form.get('name')
        batch = request.form.get('batch')
        department = request.form.get('department')
        image = request.files.get('image')
        
        if image:
            filename = secure_filename(f"{register_no}_{name}.jpg")
            image_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            image.save(image_path)
            
            student = Student(register_no=register_no,
                            name=name,
                            batch=batch,
                            department=department,
                            image_path=image_path)
            db.session.add(student)
            db.session.commit()
            flash('Student registered successfully')
            return redirect(url_for('dashboard'))
    return render_template('register_student.html')

@app.route('/attendance')
def attendance():
    if not session.get('logged_in'):
        flash('Please login first', 'warning')
        return redirect(url_for('login'))
    return render_template('attendance.html')

@app.route('/process_attendance', methods=['POST'])
def process_attendance():
    """Redirect to mark_attendance route to handle the attendance processing"""
    # This route exists to match the endpoint called from the frontend
    return mark_attendance()

@app.route('/detect_face', methods=['POST'])
def detect_face():
    """Detect if a face is present in the uploaded image"""
    if 'image' not in request.files:
        return jsonify({'error': 'No image provided', 'face_detected': False}), 400
        
    try:
        # Get the image from request
        image_file = request.files['image']
        img = cv2.imdecode(np.frombuffer(image_file.read(), np.uint8), cv2.IMREAD_COLOR)
        
        # Convert BGR to RGB for face detection
        gray_img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        rgb_img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        
        # Detect faces using MTCNN
        faces = detector.detect_faces(rgb_img)
        
        # Return face detection result
        return jsonify({
            'face_detected': len(faces) > 0,
            'confidence': faces[0]['confidence'] if faces else 0
        })
    except Exception as e:
        app.logger.error(f"Face detection error: {str(e)}")
        return jsonify({'error': f'Face detection failed: {str(e)}', 'face_detected': False}), 500

# Initialize Haar Cascade for faster face detection
face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')

@app.route('/mark_attendance', methods=['POST'])
def mark_attendance():
    if 'image' not in request.files:
        return jsonify({'error': 'No image provided'}), 400
        
    try:
        # Track total processing time
        total_start_time = time.time()
        
        # Get the image from request
        image_file = request.files['image']
        img = cv2.imdecode(np.frombuffer(image_file.read(), np.uint8), cv2.IMREAD_COLOR)
        
        # Keep higher resolution for better accuracy but don't go too high to maintain performance
        img = cv2.resize(img, (720, 540))  # Slightly higher resolution for better recognition

        # Apply noise reduction to improve face detection
        img = cv2.fastNlMeansDenoisingColored(img, None, 5, 5, 7, 21)
        
        # Convert BGR to RGB for face detection
        gray_img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        rgb_img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        
        # Apply CLAHE to improve contrast in grayscale image
        clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
        enhanced_gray = clahe.apply(gray_img)
        
        # Detect faces using MTCNN with enhanced preprocessing
        face_detection_start = time.time()
        faces = detector.detect_faces(rgb_img)
        
        # Try detection on enhanced image if no faces found
        if not faces:
            enhanced_rgb = cv2.cvtColor(enhanced_gray, cv2.COLOR_GRAY2RGB)
            faces = detector.detect_faces(enhanced_rgb)
        
        face_img = None
        face_detection_method = "none"
        
        if faces and max(faces, key=lambda x: x['confidence'])['confidence'] > 0.85:  # Slightly more permissive
            # Get the face with highest confidence from MTCNN
            face = max(faces, key=lambda x: x['confidence'])
            x, y, w, h = face['box']
            
            # Add adaptive padding around face based on face size and position
            face_size = max(w, h)
            
            # More padding for faces near the edge of the image
            edge_distance = min(x, y, img.shape[1]-x-w, img.shape[0]-y-h)
            edge_factor = max(0.1, min(0.3, 1.0 - (edge_distance / face_size)))
            
            # Base padding ratio depends on face size
            base_padding_ratio = 0.35 if face_size < 100 else 0.25
            padding_ratio = base_padding_ratio + edge_factor
            
            padding = int(padding_ratio * face_size)
            
            x_start = max(0, x - padding)
            y_start = max(0, y - padding)
            x_end = min(img.shape[1], x + w + padding)
            y_end = min(img.shape[0], y + h + padding)
            face_img = img[y_start:y_end, x_start:x_end]
            face_detection_method = "mtcnn"
            
            # Extract facial landmarks for alignment if available
            landmarks = face.get('keypoints', None)
            if landmarks and 'left_eye' in landmarks and 'right_eye' in landmarks:
                left_eye = landmarks['left_eye']
                right_eye = landmarks['right_eye']
                
                # Calculate angle for alignment
                dx = right_eye[0] - left_eye[0]
                dy = right_eye[1] - left_eye[1]
                angle = np.degrees(np.arctan2(dy, dx))
                
                # Adjust coordinates for the cropped face image
                left_eye = (left_eye[0] - x_start, left_eye[1] - y_start)
                right_eye = (right_eye[0] - x_start, right_eye[1] - y_start)
                
                # Calculate center of the face for rotation with robust error handling
                try:
                    # Convert to float first for accurate division, then to int for pixel coordinates
                    center_x = int((left_eye[0] + right_eye[0]) / 2)
                    center_y = int((left_eye[1] + right_eye[1]) / 2)
                    center = (center_x, center_y)
                    
                    # Validate center coordinates are within image bounds
                    if center_x < 0 or center_x >= face_img.shape[1] or center_y < 0 or center_y >= face_img.shape[0]:
                        app.logger.warning(f"Center coordinates {center} are outside image bounds {face_img.shape}")
                        raise ValueError("Invalid center coordinates")
                    
                    # Get rotation matrix and apply rotation
                    M = cv2.getRotationMatrix2D(center, angle, 1.0)
                    face_img = cv2.warpAffine(face_img, M, (face_img.shape[1], face_img.shape[0]), 
                                             flags=cv2.INTER_CUBIC, borderMode=cv2.BORDER_REPLICATE)
                except Exception as e:
                    app.logger.warning(f"Face alignment skipped: {str(e)}")
                    # Continue without alignment if there's an error
        else:
            # Fallback to Haar Cascade if MTCNN fails or confidence is low
            faces_haar = face_cascade.detectMultiScale(enhanced_gray, scaleFactor=1.1, minNeighbors=4, minSize=(30, 30))
            if len(faces_haar) == 0:
                # Try with original grayscale as last resort
                faces_haar = face_cascade.detectMultiScale(gray_img, scaleFactor=1.1, minNeighbors=3, minSize=(30, 30))
                if len(faces_haar) == 0:
                    face_detection_time = time.time() - face_detection_start
                    app.logger.warning(f"Face detection failed after {face_detection_time:.3f} seconds")
                    return jsonify({'error': 'No face detected in the image. Please position your face properly and ensure good lighting.'}), 400
                
            # Get the largest face from Haar Cascade
            x, y, w, h = max(faces_haar, key=lambda face: face[2] * face[3])
            
            # Add adaptive padding around face
            face_size = max(w, h)
            padding_ratio = 0.35 if face_size < 100 else 0.25  # More padding for smaller faces
            padding = int(padding_ratio * face_size)
            
            x_start = max(0, x - padding)
            y_start = max(0, y - padding)
            x_end = min(img.shape[1], x + w + padding)
            y_end = min(img.shape[0], y + h + padding)
            face_img = img[y_start:y_end, x_start:x_end]
            face_detection_method = "haar"
        
        face_detection_time = time.time() - face_detection_start
        app.logger.info(f"Face detection took {face_detection_time:.3f} seconds using {face_detection_method}")
        
        # Preprocess face image for the model
        face_img = cv2.resize(face_img, (224, 224), interpolation=cv2.INTER_LANCZOS4)
        face_img = cv2.cvtColor(face_img, cv2.COLOR_BGR2RGB)
        
        # Initialize face recognition model with caching and improved features
        model = FaceRecognitionModel()
        model.build_model()  # Build the optimized CNN model with ResNet50V2
        
        # Extract features from the captured face with improved preprocessing and alignment
        preprocess_start_time = time.time()
        try:
            # Apply enhanced preprocessing before feature extraction
            face_img = model.preprocess_image(face_img)
            preprocess_time = time.time() - preprocess_start_time
            app.logger.info(f"Face preprocessing took {preprocess_time:.3f} seconds")
        except Exception as e:
            app.logger.error(f"Face preprocessing error: {str(e)}")
            return jsonify({'error': f'Face recognition failed: {str(e)}'}), 400
        
        if face_img is None:
            return jsonify({'error': 'Failed to preprocess face image. Please try again with better lighting.'}), 400
            
        feature_start_time = time.time()
        face_embedding = model.extract_features(face_img)
        feature_time = time.time() - feature_start_time
        app.logger.info(f"Feature extraction took {feature_time:.3f} seconds")
        
        if face_embedding is None:
            app.logger.error("Failed to extract features from captured face")
            return jsonify({'error': 'Failed to extract features from face. Please try again with better lighting.'}), 400
        
        # Get all students from database
        students = Student.query.all()
        if not students:
            return jsonify({'error': 'No students in database'}), 400
        
        # Load and process all student images in batch with caching
        student_images = []
        valid_students = []
        
        # Create a cache directory if it doesn't exist
        cache_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'cache')
        if not os.path.exists(cache_dir):
            os.makedirs(cache_dir)
        
        for student in students:
            if os.path.exists(student.image_path):
                img = cv2.imread(student.image_path)
                if img is not None:
                    student_images.append(img)
                    valid_students.append(student)
        
        if not student_images:
            return jsonify({'error': 'No valid student images found'}), 400
        
        # Extract features from all student images with enhanced parallel processing and caching
        batch_start_time = time.time()
        student_embeddings, valid_indices = model.extract_features_batch(student_images)
        batch_time = time.time() - batch_start_time
        app.logger.info(f"Batch processing took {batch_time:.3f} seconds for {len(student_images)} images")
        
        if student_embeddings is None or len(valid_indices) == 0:
            return jsonify({'error': 'Failed to process student images'}), 500
            
        # Map student IDs to valid indices
        valid_students = [valid_students[i] for i in valid_indices]
        
        # Find best match using improved similarity metrics with error handling
        try:
            match_start_time = time.time()
            # Use the improved find_best_match method for more accurate matching
            student_ids = [student.id for student in valid_students]
            best_match_id, best_match_similarity = model.find_best_match(face_embedding, student_embeddings, student_ids)
            match_time = time.time() - match_start_time
            app.logger.info(f"Matching took {match_time:.3f} seconds")
            
            # Find the index of the matched student in our valid_students list
            best_match_idx = next((i for i, student in enumerate(valid_students) if student.id == best_match_id), 0)
            
            # Convert similarity to cosine similarity format for compatibility with existing code
            # (our find_best_match returns positive similarity where higher is better, but the existing code expects negative values)
            best_match_similarity = -best_match_similarity
            
            # Calculate confidence score (convert from cosine similarity to a 0-1 scale)
            # Higher negative value means more similar in cosine similarity
            confidence_score = min(1.0, max(0.0, (-best_match_similarity + 0.5) / 0.5))
            
            app.logger.info(f"Best match similarity: {best_match_similarity}, confidence: {confidence_score:.2f}")
            
            # Enhanced adaptive threshold calculation
            threshold_start_time = time.time()
            
            # Start with a more permissive base threshold
            base_threshold = -0.58  # More permissive threshold for better recall
            
            # Calculate adaptive factors with improved scaling
            # 1. Student count factor - more students means slightly more lenient threshold
            student_count_factor = min(0.06, len(valid_students) * 0.006)
            
            # 2. Time-based factor with more granular time periods
            current_hour = datetime.now().hour
            time_factor = 0
            if current_hour < 7:  # Very early morning (poor lighting)
                time_factor = 0.03
            elif current_hour < 9:  # Morning (moderate lighting)
                time_factor = 0.02
            elif current_hour > 19:  # Night (poor lighting)
                time_factor = 0.03
            elif current_hour > 17:  # Evening (moderate lighting)
                time_factor = 0.02
            
            # 3. Dynamic lighting factor based on image brightness
            lighting_factor = 0
            try:
                # Calculate average brightness and contrast of face image
                if face_img is not None and len(face_img.shape) == 3:
                    # Convert to grayscale and calculate brightness metrics
                    if isinstance(face_img, np.ndarray):
                        gray = cv2.cvtColor(face_img, cv2.COLOR_RGB2GRAY) if face_img.shape[2] == 3 else face_img
                        avg_brightness = np.mean(gray)
                        contrast = np.std(gray)
                        
                        # Adjust threshold based on brightness and contrast
                        if avg_brightness < 40 or avg_brightness > 220:  # Extreme lighting
                            lighting_factor = 0.03
                        elif avg_brightness < 60 or avg_brightness > 200:  # Poor lighting
                            lighting_factor = 0.02
                        
                        # Add contrast factor
                        if contrast < 30:  # Low contrast
                            lighting_factor += 0.02
            except Exception as e:
                app.logger.warning(f"Could not calculate image quality metrics: {str(e)}")
                lighting_factor = 0.01  # Default to slight adjustment
            
            # 4. Detection method factor - Haar cascade is less reliable than MTCNN
            detection_factor = 0.02 if face_detection_method == "haar" else 0
            
            # Combine all factors
            adaptive_threshold = base_threshold - student_count_factor - time_factor - lighting_factor - detection_factor
            
            # Log all threshold components for debugging
            threshold_info = {
                "base": base_threshold,
                "students": student_count_factor,
                "time": time_factor,
                "lighting": lighting_factor,
                "detection": detection_factor,
                "final": adaptive_threshold
            }
            app.logger.info(f"Threshold calculation: {threshold_info}")
            
            threshold_time = time.time() - threshold_start_time
            app.logger.info(f"Threshold calculation took {threshold_time:.3f} seconds")
            
            # Check if the best match exceeds our adaptive threshold
            if best_match_similarity < adaptive_threshold:  # Adaptive threshold for better matching
                matched_student = valid_students[best_match_idx]
                
                # Check if attendance already marked for today with precise timestamp
                today = datetime.now().date()
                now = datetime.now()
                
                # Get existing attendance record if any
                existing_attendance = Attendance.query.filter_by(
                    student_id=matched_student.id,
                    date=today
                ).first()
                
                if existing_attendance:
                    # Update the timestamp to the latest check-in time and ensure status is 'present'
                    existing_attendance.time = now.time()
                    existing_attendance.status = 'present'
                    existing_attendance.face_recognized = True
                    db.session.commit()
                    
                    # Calculate attendance percentage for this student
                    total_days = Attendance.query.with_entities(Attendance.date).distinct().count()
                    student_present_days = Attendance.query.filter_by(
                        student_id=matched_student.id, 
                        status='present'
                    ).count()
                    
                    attendance_percentage = 0
                    if total_days > 0:
                        attendance_percentage = (student_present_days / total_days) * 100
                    
                    return jsonify({
                        'message': f'Attendance updated for {matched_student.name} at {now.strftime("%H:%M:%S")}',
                        'student': {
                            'name': matched_student.name,
                            'register_no': matched_student.register_no,
                            'confidence': f'{confidence_score:.2f}',
                            'attendance_percentage': f'{attendance_percentage:.2f}%'
                        },
                        'success': True,
                        'timestamp': now.strftime("%Y-%m-%d %H:%M:%S"),
                        'status': 'present'
                    }), 200
                
                # Mark new attendance with timestamp
                attendance = Attendance(
                    student_id=matched_student.id,
                    date=now.date(),
                    time=now.time(),
                    status='present',
                    face_recognized=True
                )
                db.session.add(attendance)
                db.session.commit()
                
                # Update any existing 'absent' status for today to 'present'
                absent_record = Attendance.query.filter_by(
                    student_id=matched_student.id,
                    date=now.date(),
                    status='absent'
                ).first()
                
                if absent_record:
                    absent_record.status = 'present'
                    absent_record.time = now.time()
                    absent_record.face_recognized = True
                    db.session.commit()
                
                # Calculate attendance percentage for this student
                total_days = Attendance.query.with_entities(Attendance.date).distinct().count()
                student_present_days = Attendance.query.filter_by(
                    student_id=matched_student.id, 
                    status='present'
                ).count()
                
                attendance_percentage = 0
                if total_days > 0:
                    attendance_percentage = (student_present_days / total_days) * 100
                
                return jsonify({
                    'message': f'Attendance marked successfully for {matched_student.name} at {now.strftime("%H:%M:%S")}',
                    'student': {
                        'name': matched_student.name,
                        'register_no': matched_student.register_no,
                        'confidence': f'{confidence_score:.2f}',
                        'attendance_percentage': f'{attendance_percentage:.2f}%'
                    },
                    'success': True,
                    'timestamp': now.strftime("%Y-%m-%d %H:%M:%S"),
                    'status': 'present'
                }), 200
        except Exception as e:
            app.logger.error(f"Error during similarity calculation: {str(e)}")
            return jsonify({'error': f'Face matching error: {str(e)}'}), 500
        else:
            # Even if confidence is low, find the best possible match for approximate matching
            # Get the best match regardless of threshold for approximate matching
            matched_student = valid_students[best_match_idx]
            
            # Log that we're using approximate matching with lower confidence
            app.logger.info(f"Using approximate matching with lower confidence: {-best_match_similarity:.2f}")
            
            # Check if attendance already marked for today
            today = datetime.now().date()
            now = datetime.now()
            
            # Get existing attendance record if any
            existing_attendance = Attendance.query.filter_by(
                student_id=matched_student.id,
                date=today
            ).first()
            
            if existing_attendance:
                # Update the timestamp to the latest check-in time and ensure status is 'present'
                existing_attendance.time = now.time()
                existing_attendance.status = 'present'
                existing_attendance.face_recognized = True
                db.session.commit()
                
                # Calculate attendance percentage for this student
                total_days = Attendance.query.with_entities(Attendance.date).distinct().count()
                student_present_days = Attendance.query.filter_by(
                    student_id=matched_student.id, 
                    status='present'
                ).count()
                
                attendance_percentage = 0
                if total_days > 0:
                    attendance_percentage = (student_present_days / total_days) * 100
                
                return jsonify({
                    'message': f'Attendance updated for {matched_student.name} at {now.strftime("%H:%M:%S")} (approximate match)',
                    'student': {
                        'name': matched_student.name,
                        'register_no': matched_student.register_no,
                        'confidence': f'{confidence_score:.2f} (approximate)',
                        'attendance_percentage': f'{attendance_percentage:.2f}%'
                    },
                    'success': True,
                    'timestamp': now.strftime("%Y-%m-%d %H:%M:%S"),
                    'status': 'present',
                    'approximate_match': True
                }), 200
            
            # Mark new attendance with timestamp
            attendance = Attendance(
                student_id=matched_student.id,
                date=now.date(),
                time=now.time(),
                status='present',
                face_recognized=True
            )
            db.session.add(attendance)
            db.session.commit()
            
            # Update any existing 'absent' status for today to 'present'
            absent_record = Attendance.query.filter_by(
                student_id=matched_student.id,
                date=now.date(),
                status='absent'
            ).first()
            
            if absent_record:
                absent_record.status = 'present'
                absent_record.time = now.time()
                absent_record.face_recognized = True
                db.session.commit()
            
            # Calculate attendance percentage for this student
            total_days = Attendance.query.with_entities(Attendance.date).distinct().count()
            student_present_days = Attendance.query.filter_by(
                student_id=matched_student.id, 
                status='present'
            ).count()
            
            attendance_percentage = 0
            if total_days > 0:
                attendance_percentage = (student_present_days / total_days) * 100
            
            return jsonify({
                'message': f'Attendance marked successfully for {matched_student.name} at {now.strftime("%H:%M:%S")} (approximate match)',
                'student': {
                    'name': matched_student.name,
                    'register_no': matched_student.register_no,
                    'confidence': f'{confidence_score:.2f} (approximate)',
                    'attendance_percentage': f'{attendance_percentage:.2f}%'
                },
                'success': True,
                'timestamp': now.strftime("%Y-%m-%d %H:%M:%S"),
                'status': 'present',
                'approximate_match': True
            }), 200
            
    except Exception as e:
        app.logger.error(f"Face recognition error: {str(e)}")
        return jsonify({'error': f'Face recognition failed: {str(e)}'}), 500

@app.route('/student_attendance_history/<int:student_id>')
def student_attendance_history(student_id):
    if not session.get('logged_in'):
        return jsonify({'error': 'Authentication required'}), 401
        
    # Get the student
    student = Student.query.get(student_id)
    if not student:
        return jsonify({'error': 'Student not found'}), 404
        
    # Get all attendance records for this student
    attendance_records = Attendance.query.filter_by(student_id=student_id).order_by(Attendance.date.desc(), Attendance.time.desc()).all()
    
    # Format the attendance records with timestamps
    attendance_data = [{
        'date': record.date.strftime('%Y-%m-%d'),
        'time': record.time.strftime('%H:%M:%S') if record.time else None,
        'status': record.status,
        'face_recognized': record.face_recognized
    } for record in attendance_records]
    
    return jsonify({
        'student': {
            'id': student.id,
            'register_no': student.register_no,
            'name': student.name,
            'batch': student.batch,
            'department': student.department
        },
        'attendance': attendance_data
    })

@app.route('/delete_student/<int:student_id>', methods=['POST'])
def delete_student(student_id):
    if not session.get('logged_in'):
        flash('Please login first', 'warning')
        return redirect(url_for('login'))
    
    student = Student.query.get(student_id)
    if not student:
        flash('Student not found', 'danger')
        return redirect(url_for('dashboard'))
    
    try:
        # Delete the student's image file
        if os.path.exists(student.image_path):
            os.remove(student.image_path)
        
        # Delete the student's attendance records
        Attendance.query.filter_by(student_id=student_id).delete()
        
        # Delete the student
        db.session.delete(student)
        db.session.commit()
        
        flash('Student deleted successfully', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting student: {str(e)}', 'danger')
    
    return redirect(url_for('dashboard'))

@app.route('/download_report')
def download_report():
    if not session.get('logged_in'):
        flash('Please login first', 'warning')
        return redirect(url_for('login'))
        
    try:
        # Get report type from query parameters (default to Excel)
        report_format = request.args.get('format', 'excel')
        
        # Get all students
        students = Student.query.all()
        
        # Get today's date for status calculation
        today = datetime.now().date()
        
        # Prepare data for the report
        report_data = []
        for student in students:
            # Get all attendance records for this student
            attendance_records = Attendance.query.filter_by(student_id=student.id).order_by(Attendance.date.desc()).all()
            
            # Determine current status
            status = 'Absent'
            last_attendance_date = None
            if attendance_records:
                last_record = attendance_records[0]
                last_attendance_date = last_record.date
                if last_record.date == today:
                    status = 'Present'
            
            # Calculate attendance percentage
            total_days = Attendance.query.with_entities(Attendance.date).distinct().count()
            student_present_days = Attendance.query.filter_by(
                student_id=student.id, 
                status='present'
            ).count()
            
            attendance_percentage = 0
            if total_days > 0:
                attendance_percentage = (student_present_days / total_days) * 100
            
            # Get last attendance time if available
            last_attendance_time = None
            if attendance_records and attendance_records[0].time:
                last_attendance_time = attendance_records[0].time
                
            # Add to report data
            report_data.append({
                'Register No': student.register_no,
                'Name': student.name,
                'Batch': student.batch,
                'Department': student.department,
                'Last Attendance Date': last_attendance_date.strftime('%Y-%m-%d') if last_attendance_date else 'Never',
                'Last Attendance Time': last_attendance_time.strftime('%H:%M:%S') if last_attendance_time else 'N/A',
                'Current Status': status,
                'Attendance Percentage': f'{attendance_percentage:.2f}%',
                'Present Days': student_present_days,
                'Total Days': total_days
            })
        
        # Create DataFrame
        df = pd.DataFrame(report_data)
        
        # Generate file based on requested format
        if report_format.lower() == 'csv':
            # Create CSV file
            with tempfile.NamedTemporaryFile(suffix='.csv', delete=False) as tmp:
                df.to_csv(tmp.name, index=False)
                return send_file(
                    tmp.name,
                    mimetype='text/csv',
                    as_attachment=True,
                    download_name=f'attendance_report_{today.strftime("%Y%m%d")}.csv'
                )
        else:
            # Create Excel file with formatting
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                with pd.ExcelWriter(tmp.name, engine='openpyxl') as writer:
                    # Write main attendance sheet
                    df.to_excel(writer, index=False, sheet_name='Attendance Summary')
                    worksheet = writer.sheets['Attendance Summary']
                    
                    # Format headers
                    for col in range(len(df.columns)):
                        cell = worksheet.cell(row=1, column=col + 1)
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
                    
                    # Format status cells
                    for row in range(2, len(df) + 2):
                        status_cell = worksheet.cell(row=row, column=df.columns.get_loc('Current Status') + 1)
                        if status_cell.value == 'Present':
                            status_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                        else:
                            status_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                    
                    # Auto-adjust column widths
                    for column in worksheet.columns:
                        max_length = 0
                        column = [cell for cell in column]
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = (max_length + 2)
                        worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
                    
                    # Add department-wise summary sheet
                    dept_summary = df.groupby('Department').agg({
                        'Register No': 'count',
                        'Present Days': 'sum',
                        'Total Days': 'mean'
                    }).reset_index()
                    
                    dept_summary.columns = ['Department', 'Student Count', 'Total Present Days', 'Total Days']
                    dept_summary['Attendance Percentage'] = (dept_summary['Total Present Days'] / 
                                                          (dept_summary['Student Count'] * dept_summary['Total Days']) * 100)
                    dept_summary['Attendance Percentage'] = dept_summary['Attendance Percentage'].apply(lambda x: f'{x:.2f}%')
                    
                    dept_summary.to_excel(writer, index=False, sheet_name='Department Summary')
                    dept_sheet = writer.sheets['Department Summary']
                    
                    # Format department summary headers
                    for col in range(len(dept_summary.columns)):
                        cell = dept_sheet.cell(row=1, column=col + 1)
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
                    
                    # Auto-adjust department summary column widths
                    for column in dept_sheet.columns:
                        max_length = 0
                        column = [cell for cell in column]
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = (max_length + 2)
                        dept_sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
                
                # Return the file
                return send_file(
                    tmp.name,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    download_name=f'attendance_report_{today.strftime("%Y%m%d")}.xlsx'
                )
    
    except Exception as e:
        app.logger.error(f'Error generating report: {str(e)}')
        flash(f'Error generating report: {str(e)}', 'danger')
        return redirect(url_for('dashboard'))